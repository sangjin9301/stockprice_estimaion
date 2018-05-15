from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelWriter:

    def __init__(self):
        self.path = 'D:\\stockprice_estimaion\\Evaluation\\'
        self.file_name = 'result_moreData.xlsx'
        self.row_index = 1

    def generateFile(self):
        wb = Workbook()
        ws = wb.create_sheet('train', 0)
        ws = wb.create_sheet('evaluation', 1)

        wb.save(self.path+self.file_name)

    def write_train_data(self, data):
        wb = load_workbook(self.path+self.file_name)
        wb.active = 0
        ws = wb.active

        ws['A' + str(self.row_index)] = self.row_index
        ws['B' + str(self.row_index)] = data

        wb.save(self.path+self.file_name)
        self.row_index += 1

    def write_evaluation_data(self, data):
        wb = load_workbook(self.path+self.file_name)
        wb.active = 1
        ws = wb.active

        ws['A' + str(self.row_index)] = self.row_index
        ws['B' + str(self.row_index)] = data

        wb.save(self.path+self.file_name)
        self.row_index += 1
